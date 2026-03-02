from fastapi import APIRouter, File, Form, Header, HTTPException, UploadFile
from pydantic import BaseModel
from supabase import create_client
from app.config import SUPABASE_URL, SUPABASE_SERVICE_KEY
from app.services.parser import parse_questionnaire
from app.services.ai_service import answer_question
from docx import Document
from pdfminer.high_level import extract_text
from openpyxl import load_workbook
import csv
import io

router = APIRouter()

def get_supabase(token: str):
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

class QuestionnaireSubmit(BaseModel):
    title: str
    content: str

class AnswerUpdate(BaseModel):
    answer: str


def extract_questionnaire_content(filename: str, data: bytes) -> str:
    lower_name = (filename or "").lower()

    if lower_name.endswith((".txt", ".md")):
        return data.decode("utf-8", errors="ignore").strip()

    if lower_name.endswith(".docx"):
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs).strip()

    if lower_name.endswith(".pdf"):
        return extract_text(io.BytesIO(data)).strip()

    if lower_name.endswith(".csv"):
        rows = []
        decoded = data.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(decoded))
        for row in reader:
            text = " ".join(str(c).strip() for c in row if str(c).strip())
            if text:
                rows.append(text)
        return "\n".join(rows).strip()

    if lower_name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(data), data_only=True)
        lines = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if values:
                    lines.append(" ".join(values))
        return "\n".join(lines).strip()

    raise HTTPException(
        status_code=400,
        detail="Unsupported questionnaire format. Use .txt, .md, .docx, .pdf, .csv, or .xlsx"
    )


def process_questionnaire_run(supabase, user_id: str, title: str, content: str) -> str:
    # Parse questions
    questions = parse_questionnaire(content)
    if not questions:
        raise HTTPException(status_code=400, detail="No questions found in questionnaire")

    # Get reference documents
    docs_result = supabase.table("documents").select("*").eq("user_id", user_id).execute()
    reference_docs = docs_result.data

    if not reference_docs:
        raise HTTPException(status_code=400, detail="No reference documents found. Please upload documents first.")

    # Create questionnaire run
    run_result = supabase.table("questionnaire_runs").insert({
        "user_id": user_id,
        "title": title,
        "status": "processing",
        "coverage_total": len(questions)
    }).execute()

    run_id = run_result.data[0]["id"]

    # Answer each question
    answered = 0
    not_found = 0

    for q in questions:
        try:
            result = answer_question(q["question_text"], reference_docs)
        except Exception as e:
            # Expose a concise upstream failure so frontend can display the real cause.
            raise HTTPException(
                status_code=502,
                detail=f"AI processing failed: {str(e)}"
            )

        if "Not found in references" in result["answer"]:
            not_found += 1
        else:
            answered += 1

        supabase.table("questions").insert({
            "run_id": run_id,
            "question_number": q["question_number"],
            "question_text": q["question_text"],
            "answer": result["answer"],
            "citations": result["citations"],
            "confidence": result["confidence"]
        }).execute()

    # Update coverage summary
    supabase.table("questionnaire_runs").update({
        "status": "completed",
        "coverage_answered": answered,
        "coverage_not_found": not_found
    }).eq("id", run_id).execute()

    return run_id


@router.post("/run")
async def run_questionnaire(payload: QuestionnaireSubmit, authorization: str = Header(...)):
    token = authorization.replace("Bearer ", "")
    supabase = get_supabase(token)

    user = supabase.auth.get_user(token)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    user_id = user.user.id
    run_id = process_questionnaire_run(supabase, user_id, payload.title, payload.content)
    return {"message": "Questionnaire processed", "run_id": run_id}


@router.post("/run-file")
async def run_questionnaire_file(
    file: UploadFile = File(...),
    title: str | None = Form(default=None),
    authorization: str = Header(...)
):
    token = authorization.replace("Bearer ", "")
    supabase = get_supabase(token)

    user = supabase.auth.get_user(token)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    user_id = user.user.id

    file_data = await file.read()
    if not file_data:
        raise HTTPException(status_code=400, detail="Uploaded questionnaire file is empty")

    content = extract_questionnaire_content(file.filename or "", file_data)
    if not content:
        raise HTTPException(status_code=400, detail="No readable content found in questionnaire file")

    run_title = (title or (file.filename or "Questionnaire Run")).strip()
    run_id = process_questionnaire_run(supabase, user_id, run_title, content)

    return {"message": "Questionnaire file processed", "run_id": run_id}

@router.get("/runs")
async def get_runs(authorization: str = Header(...)):
    token = authorization.replace("Bearer ", "")
    supabase = get_supabase(token)

    user = supabase.auth.get_user(token)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    user_id = user.user.id
    result = supabase.table("questionnaire_runs").select("*").eq("user_id", user_id).order("created_at", desc=True).execute()
    return {"runs": result.data}

@router.get("/{run_id}/questions")
async def get_questions(run_id: str, authorization: str = Header(...)):
    token = authorization.replace("Bearer ", "")
    supabase = get_supabase(token)

    result = supabase.table("questions").select("*").eq("run_id", run_id).order("question_number").execute()
    return {"questions": result.data}

@router.patch("/{question_id}/edit")
async def edit_answer(question_id: str, payload: AnswerUpdate, authorization: str = Header(...)):
    token = authorization.replace("Bearer ", "")
    supabase = get_supabase(token)

    supabase.table("questions").update({
        "answer": payload.answer,
        "is_edited": True
    }).eq("id", question_id).execute()

    return {"message": "Answer updated"}


@router.delete("/{run_id}")
async def delete_run(run_id: str, authorization: str = Header(...)):
    token = authorization.replace("Bearer ", "")
    supabase = get_supabase(token)

    user = supabase.auth.get_user(token)
    if not user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    user_id = user.user.id

    run_result = (
        supabase.table("questionnaire_runs")
        .select("id")
        .eq("id", run_id)
        .eq("user_id", user_id)
        .execute()
    )

    if not run_result.data:
        raise HTTPException(status_code=404, detail="Run not found")

    supabase.table("questions").delete().eq("run_id", run_id).execute()
    supabase.table("questionnaire_runs").delete().eq("id", run_id).eq("user_id", user_id).execute()

    return {"message": "Run deleted"}
